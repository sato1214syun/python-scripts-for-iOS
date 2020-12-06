import os
import glob
import sys
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

import dropbox
import pandas as pd
import requests as rq
from bs4 import BeautifulSoup as bs
import urllib.parse
import datetime as dt
from time import sleep
from tqdm import tqdm
from pprint import pprint


#dropboxアクセス用の情報
TOKEN = ""
fileName = "交通費計算.xlsx"
tempFileName = "temp_{}".format(fileName)
tempFilePath = "./{}".format(tempFileName)
DBXPath = "/家族共有フォルダ/愛 経費明細/{}".format(fileName)

#dropboxからファイルをダウンロード
if not os.path.exists(tempFileName):
    dbx = dropbox.Dropbox(TOKEN)
    dbx.users_get_current_account()
    dbx.files_download_to_file(tempFilePath, DBXPath)


#ダウンロードしたエクセルファイルから情報を抽出
wb = openpyxl.load_workbook(filename=tempFileName, read_only=False)
sheetNames = wb.sheetnames
print("シート名一覧:{}".format(sheetNames))
print("運賃検索を行うシート名を入力してください")
while(True):
    sheetName = input()
    if sheetName in sheetNames:

        break
    else:
        print("存在しないシート名です。もう一度入力してください")

print("運賃を検索しています...")
sheet = wb[sheetName]
currentData = sheet.values
df = pd.DataFrame(currentData)
df = df.iloc[4:,1:8]
df = df.dropna(how="all")
df.columns = df.iloc[0].values.tolist()
df = df.iloc[1:]
df = df.reset_index(drop=True)

# 運賃を取得
url = "https://transit.yahoo.co.jp/search/result"
origParams = {
    "from": "",
    "to": "",
    "via": "",
    "y": "",
    "m": "",
    "d": "",
    "hh": "",
    "m1": "",
    "m2": "",
    "type": "1",
    "ticket": "ic",
    "expkind": "1",
    "ws": "2",
    "s": "0",
    "al": "0",
    "shin": "0",
    "ex": "0",
    "hb": "0",
    "lb": "0",
    "sr": "0",
    "kw": ""
}

for data in tqdm(df.itertuples(), total=len(df.index)):
    if data[5] == None:
        timeSetting = data[1].replace(hour=8, minute=0, second=0)
        departure = data[2]
        destination = data[3]
        via = ""
        params = origParams
        updateParams = {
            "from": departure,
            "to": destination,
            "via": via,
            "y": timeSetting.strftime("%Y"),
            "m": timeSetting.strftime("%m"),
            "d": timeSetting.strftime("%d"),
            "hh": timeSetting.strftime("%H"),
            "m1": timeSetting.strftime("%M")[:1],
            "m2": timeSetting.strftime("%M")[-1:],
            "kw": destination
        }
        params.update(updateParams)
        rs = rq.get(url, params=params)
        soup = bs(rs.content, "lxml")
        trainFee = int(soup.find("li", class_="fare").string[:-1].replace(",", ""))
        if data[4] == "片道":
            trainFee *= 1
        elif data[4] == "往復":
            trainFee *= 2
        else:
            trainFee =0
        df.at[data[0], "運賃 [円]"] = trainFee
        df.at[data[0], "URL"] = rs.url
        #sleep(0.2)

print("運賃検索を完了しました")

#エクセルファイルに運賃を書き込み
rows = openpyxl.utils.dataframe.dataframe_to_rows(df, index=False, header=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
         sheet.cell(row=r_idx+5, column=c_idx+1, value=value)
wb.save(filename=fileName)

'''
uploadFilePath = "./{}".format(fileName)

if not dbx:
    dbx = dropbox.Dropbox(TOKEN)
    dbx.users_get_current_account()

with open(uploadFilePath, "rb") as f:
    dbx.files_upload(f.read(), DBXPath, mode=dropbox.files.WriteMode.overwrite)

deleteFileList = glob.glob("*交通費計算.xlsm")
for file in deleteFileList:
    print("remove：{0}".format(file))
    os.remove(file)
'''
print("Completed")
